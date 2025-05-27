"""This module contains the main process of the robot."""

# import sys

import json
import urllib.parse

import datetime

from typing import Dict, Any

from io import BytesIO

import math

import pandas as pd

import openpyxl

from sqlalchemy import create_engine

from openpyxl.styles import Alignment

from OpenOrchestrator.orchestrator_connection.connection import OrchestratorConnection

# from itk_dev_shared_components.smtp import smtp_util

from mbu_dev_shared_components.msoffice365.sharepoint_api.files import Sharepoint

from robot_framework.helper_functions import formular_mappings

from robot_framework.helper_functions import upload_pdf_to_sharepoint

SHAREPOINT_FOLDER_URL = "https://aarhuskommune.sharepoint.com"
SHAREPOINT_DOCUMENT_LIBRARY = "Delte dokumenter"


def process(orchestrator_connection: OrchestratorConnection) -> None:
    """Do the primary process of the robot."""

    orchestrator_connection.log_trace("Running process.")
    print("Running process.")

    new_forms = []

    site_name = ""
    folder_name = ""
    excel_file_name = ""
    formular_mapping = None

    upload_pdfs_to_sharepoint = False
    upload_pdfs_to_sharepoint_folder_name = ""

    current_day_of_month = str(pd.Timestamp.now().day)

    print("Current day of month:", current_day_of_month)
    print("Current day of month type:", type(current_day_of_month))

    sql_server_connection_string = orchestrator_connection.get_constant("DbConnectionString").value

    proc_args = json.loads(orchestrator_connection.process_arguments)

    os2_webform_id = proc_args["os2_webform_id"]

    username = orchestrator_connection.get_credential("SvcRpaMBU002").username
    password = orchestrator_connection.get_credential("SvcRpaMBU002").password

    os2_api_key = orchestrator_connection.get_credential('os2_api').password

    if os2_webform_id == "basisteam_spoergeskema_til_fagpe":
        site_name = "tea-teamsite11462"
        folder_name = "General/Udtræk OS2Forms"
        excel_file_name = "Dataudtræk basisteam - fagperson.xlsx"
        formular_mapping = formular_mappings.basisteam_spoergeskema_til_fagpe_mapping

        upload_pdfs_to_sharepoint = True
        upload_pdfs_to_sharepoint_folder_name = "General/Besvarelser fra OS2Forms - fagpersoner"

    elif os2_webform_id == "basisteam_spoergeskema_til_forae":
        site_name = "tea-teamsite11462"
        folder_name = "General/Udtræk OS2Forms"
        excel_file_name = "Dataudtræk basisteam - forældre.xlsx"
        formular_mapping = formular_mappings.basisteam_spoergeskema_til_forae_mapping

        upload_pdfs_to_sharepoint = True
        upload_pdfs_to_sharepoint_folder_name = "General/Besvarelser fra OS2Forms - forældre"

    elif os2_webform_id == "henvisningsskema_til_klinisk_hyp":
        site_name = "tea-teamsite10693"
        folder_name = "General/Udtræk OS2Forms/Henvisningsskema"
        excel_file_name = "Dataudtræk henvisningsskema hypnoterapi.xlsx"
        formular_mapping = formular_mappings.henvisningsskema_til_klinisk_hyp_mapping

    elif os2_webform_id == "spoergeskema_hypnoterapi_foer_fo":
        site_name = "tea-teamsite10693"
        folder_name = "General/Udtræk OS2Forms/Spørgeskema"
        excel_file_name = "Dataudtræk spørgeskema hypnoterapi.xlsx"
        formular_mapping = formular_mappings.spoergeskema_hypnoterapi_foer_fo_mapping

    elif os2_webform_id == "opfoelgende_spoergeskema_hypnote":
        site_name = "tea-teamsite10693"
        folder_name = "General/Udtræk OS2Forms/Opfølgende spørgeskema"
        excel_file_name = "Dataudtræk opfølgende spørgeskema hypnoterapi.xlsx"
        formular_mapping = formular_mappings.opfoelgende_spoergeskema_hypnote_mapping

    elif os2_webform_id == "foraelder_en_god_overgang_fra_hj":
        site_name = "tea-teamsite10533"
        folder_name = "General/Udtræk data OS2Forms/Opfølgende spørgeskema forældre"
        excel_file_name = "Dataudtræk en god overgang fra hjem til dagtilbud - forælder.xlsx"
        formular_mapping = formular_mappings.foraelder_en_god_overgang_fra_hj_mapping

    elif os2_webform_id == "fagperson_en_god_overgang_fra_hj":
        site_name = "tea-teamsite10533"
        folder_name = "General/Udtræk data OS2Forms/Opfølgende spørgeskema fagpersonale"
        excel_file_name = "Dataudtræk en god overgang fra hjem til dagtilbud - fagperson.xlsx"
        formular_mapping = formular_mappings.fagperson_en_god_overgang_fra_hj_mapping

    elif os2_webform_id == "sundung_aarhus":
        site_name = "tea-teamsite11121"
        folder_name = "General/Udtræk OS2-formularer"
        excel_file_name = "Dataudtræk SundUng Aarhus.xlsx"
        formular_mapping = formular_mappings.sundung_aarhus_mapping

    elif os2_webform_id == "tilmelding_til_modersmaalsunderv":
        today = datetime.date.today()
        # today = datetime.date(2025, 5, 26)

        monday_last_week = today - datetime.timedelta(days=today.weekday() + 7)
        sunday_last_week = monday_last_week + datetime.timedelta(days=6)

        site_name = "Teams-Modersmlsundervisning"
        folder_name = "General"
        formular_mapping = formular_mappings.tilmelding_til_modersmaalsunderv_mapping
        excel_file_name = f"Dataudtræk - {monday_last_week} til {sunday_last_week}.xlsx"

    sharepoint_api = Sharepoint(username=username, password=password, site_url=SHAREPOINT_FOLDER_URL, site_name=site_name, document_library=SHAREPOINT_DOCUMENT_LIBRARY)

    # STEP 1 - Get all active forms from the SQL server
    orchestrator_connection.log_trace("STEP 1 - Fetching all active forms.")
    print("STEP 1 - Fetching all active forms.")
    all_active_forms = get_forms_data(sql_server_connection_string, os2_webform_id)
    orchestrator_connection.log_trace(f"OS2 forms retrieved. {len(all_active_forms)} active forms found.")
    print(f"OS2 forms retrieved. {len(all_active_forms)} active forms found.")

    # Modersmaalsundervisning has a different flow - therefore we skip the Excel overwrite functionality if we are currently running that formular
    if os2_webform_id != "tilmelding_til_modersmaalsunderv":
        if str(current_day_of_month) == "1":
            orchestrator_connection.log_trace("Current day is the first of the month - updating Excel file with new submissions.")
            print("Current day is the first of the month - updating Excel file with new submissions.")

            # STEP 2 - Get the Excel file from Sharepoint
            orchestrator_connection.log_trace("STEP 2 - Retrieving existing Excel sheet.")
            print("STEP 2 - Retrieving existing Excel sheet.")
            excel_file = sharepoint_api.fetch_file_using_open_binary(excel_file_name, folder_name)
            excel_stream = BytesIO(excel_file)
            excel_file_df = pd.read_excel(excel_stream)
            orchestrator_connection.log_trace(f"Excel file retrieved. {len(excel_file_df)} rows found in existing sheet.")
            print(f"Excel file retrieved. {len(excel_file_df)} rows found in existing sheet.")

            # Create a set of serial numbers from the Excel file
            serial_set = set(excel_file_df["Serial number"].tolist())

            # STEP 3 - Loop through all active forms and transform them to the correct format
            orchestrator_connection.log_trace("STEP 3 - Looping forms and mapping retrieved data to fit Excel column names.")
            print("STEP 3 - Looping forms and mapping retrieved data to fit Excel column names.")
            for form in all_active_forms:
                form_serial_number = form["entity"]["serial"][0]["value"]

                if form_serial_number in serial_set:
                    continue

                transformed_row = formular_mappings.transform_form_submission(form_serial_number, form, formular_mapping)

                new_forms.append(transformed_row)

            # STEP 4 & 5 - If new forms are found, append them to the Excel file, format the file and upload it to Sharepoint
            if new_forms:
                orchestrator_connection.log_trace(f"New forms found. {len(new_forms)} new forms to be added.")
                print(f"New forms found. {len(new_forms)} new forms to be added.")

                new_forms_df = pd.DataFrame(new_forms)

                # Append the new forms to the existing DataFrame
                updated_excel_df = pd.concat([excel_file_df, new_forms_df], ignore_index=True)

                # Sort by "Serial number" in descending order
                updated_excel_df.sort_values(by="Serial number", ascending=False, inplace=True)

                # Save the updated DataFrame to an in-memory Excel file
                updated_excel_stream = BytesIO()
                updated_excel_df.to_excel(updated_excel_stream, index=False, engine="openpyxl")
                updated_excel_stream.seek(0)

                # Apply formatting and get the formatted stream
                orchestrator_connection.log_trace("STEP 4 - Formatting Excel file.")
                print("STEP 4 - Formatting Excel file.")
                formatted_stream = format_excel_file(updated_excel_stream)

                # Upload the formatted Excel file to SharePoint
                orchestrator_connection.log_trace("STEP 5 - Uploading formatted Excel file to Sharepoint.")
                print("STEP 5 - Uploading formatted Excel file to Sharepoint.")
                sharepoint_api.upload_file_from_bytes(
                    binary_content=formatted_stream.getvalue(),
                    file_name=excel_file_name,
                    folder_name=folder_name
                )

            else:
                print("No new forms found.")

                orchestrator_connection.log_trace("No new forms found.")
                print("No new forms found.")

        if upload_pdfs_to_sharepoint:
            # Upload the PDFs to SharePoint
            orchestrator_connection.log_trace("Uploading PDFs to SharePoint.")
            print("Uploading PDFs to SharePoint.")

            upload_pdf_to_sharepoint.upload_pdf_to_sharepoint(
                orchestrator_connection=orchestrator_connection,
                sharepoint_api=sharepoint_api,
                folder_name=upload_pdfs_to_sharepoint_folder_name,
                os2_api_key=os2_api_key,
                active_forms=all_active_forms
            )

    # The webform is tilmelding_til_modersmaalsundervising - therefore we run a different flow
    else:
        orchestrator_connection.log_trace(f"Filtering forms completed between {monday_last_week} and {sunday_last_week}")
        print(f"Filtering forms completed between {monday_last_week} and {sunday_last_week}")

        # We start by filtering out submissions that were not submitted in the previous week
        for form in all_active_forms:
            form_serial_number = form["entity"]["serial"][0]["value"]

            completed_str = form["entity"]["completed"][0]["value"]

            if completed_str:
                completed_time = datetime.datetime.fromisoformat(completed_str).date()

                if monday_last_week <= completed_time <= sunday_last_week:
                    transformed_row = formular_mappings.transform_form_submission(form_serial_number, form, formular_mapping)

                    new_forms.append(transformed_row)

        # If any new forms were submitted for the previous week, we initialize a dataframe of these
        if new_forms:
            new_forms_df = pd.DataFrame(new_forms)

            # Sort by "Serial number" in descending order
            new_forms_df.sort_values(by="Ønsket sprog", ascending=True, inplace=True)

            # Save the updated DataFrame to an in-memory Excel file
            updated_excel_stream = BytesIO()
            new_forms_df.to_excel(updated_excel_stream, index=False, engine="openpyxl")
            updated_excel_stream.seek(0)

            # Apply formatting and get the formatted stream
            formatted_stream = format_excel_file(updated_excel_stream)

            # Upload the formatted Excel file to SharePoint
            sharepoint_api.upload_file_from_bytes(
                binary_content=formatted_stream.getvalue(),
                file_name=excel_file_name,
                folder_name=folder_name
            )

    orchestrator_connection.log_trace("Process completed successfully.")
    print("Process completed successfully.")

    return "Process completed successfully."


def format_excel_file(excel_stream: BytesIO) -> BytesIO:
    """
    Applies formatting to an Excel file contained in a BytesIO stream.
    This includes:
      - Freezing the first row.
      - Applying left and top alignment to all cells.
      - Auto-adjusting column widths up to a maximum width and enabling wrap_text if needed.
      - Auto-adjusting row heights based on the wrapped text.

    Returns:
        A new BytesIO stream containing the formatted workbook.
    """

    # Load the workbook from the input stream
    wb = openpyxl.load_workbook(excel_stream)
    ws = wb.active

    # Freeze the first row
    ws.freeze_panes = "A2"

    # Apply left alignment and top vertical alignment to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="top")

    # Define a maximum column width (in characters)
    max_allowed_width = 100  # adjust as needed

    # Auto-adjust column widths based on content length, enabling wrap_text if necessary
    for col in ws.columns:
        max_length = 0

        column_letter = col[0].column_letter  # Get column letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        computed_width = max_length + 2

        if computed_width > max_allowed_width:
            ws.column_dimensions[column_letter].width = max_allowed_width

            # Enable wrap_text for cells in this column
            for cell in col:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        else:
            ws.column_dimensions[column_letter].width = computed_width

    # Auto-adjust row heights based on wrapped text (simulate double-click auto-fit)
    for row in ws.iter_rows():
        max_line_count = 1  # Start with at least one line

        for cell in row:
            if cell.value and cell.alignment.wrap_text:
                col_letter = cell.column_letter

                # Use the set column width or a default value if not set
                col_width = ws.column_dimensions[col_letter].width or 10

                # Estimate how many characters fit in one line (factor may need tweaking)
                chars_per_line = col_width * 1.2

                # Split the cell text by newlines
                lines = str(cell.value).split("\n")

                # Estimate total line count for the cell
                line_count = sum(math.ceil(len(line) / chars_per_line) for line in lines)

                max_line_count = max(max_line_count, line_count)

        # Set the row height (multiplier of 20 is a rough estimate; adjust as needed)
        ws.row_dimensions[row[0].row].height = max_line_count * 20

    # Save the formatted workbook to a new BytesIO stream and return it
    formatted_stream = BytesIO()

    wb.save(formatted_stream)

    formatted_stream.seek(0)

    return formatted_stream


def get_forms_data(conn_string: str, form_type: str) -> list[dict]:
    """
    Retrieve form_data['data'] for all matching submissions for the given form type,
    excluding purged entries.
    """

    print("inside get_forms_data")

    query = """
        SELECT
            form_id,
            form_data,
            CAST(form_submitted_date AS datetime) AS form_submitted_date
        FROM
            [RPA].[journalizing].[Forms]
        WHERE
            form_type = ?
            AND form_data IS NOT NULL
            AND form_submitted_date IS NOT NULL
        ORDER BY form_submitted_date DESC
    """

    # Create SQLAlchemy engine
    encoded_conn_str = urllib.parse.quote_plus(conn_string)
    engine = create_engine(f"mssql+pyodbc:///?odbc_connect={encoded_conn_str}")

    try:
        df = pd.read_sql(sql=query, con=engine, params=(form_type,))

    except Exception as e:
        print("Error during pd.read_sql:", e)

        raise

    if df.empty:
        print("No submissions found for the given form type.")

        return []

    extracted_data = []
    for _, row in df.iterrows():
        try:
            parsed = json.loads(row["form_data"])
            if "purged" not in parsed:  # Skip purged entries
                extracted_data.append(parsed)

        except json.JSONDecodeError:
            print("Invalid JSON in form_data, skipping row.")

    return extracted_data


def get_credentials_and_constants(orchestrator_connection: OrchestratorConnection) -> Dict[str, Any]:
    """Retrieve necessary credentials and constants from the orchestrator connection."""
    try:
        credentials = {
            "go_api_endpoint": orchestrator_connection.get_constant('go_api_endpoint').value,
            "go_api_username": orchestrator_connection.get_credential('go_api').username,
            "go_api_password": orchestrator_connection.get_credential('go_api').password,
            "os2_api_key": orchestrator_connection.get_credential('os2_api').password,
            "sql_conn_string": orchestrator_connection.get_constant('DbConnectionString').value,
            "journalizing_tmp_path": orchestrator_connection.get_constant('journalizing_tmp_path').value,
        }
        return credentials
    except AttributeError as e:
        raise SystemExit(e) from e
